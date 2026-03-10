"""
build_subset_request_form.py
============================
Generates  Subset_Request_Form.xlsx — one sheet per study containing:
  • Study overview  (N, visits, groups, procedures)
  • Procedure participation  (# participants with values)
  • Table 1 characteristics  (age, sex, BMI, UACR, eGFR) stratified by group
  • PB90 scRNA inventory  (cell counts by cell type and group, from Kopah S3)

Run:
    /Users/shivaniramesh/.pyenv/versions/3.11.14/bin/python3.11 build_subset_request_form.py

Dependencies:
    pip install openpyxl boto3==1.26.160 botocore==1.29.160 pyreadr
"""

import csv
import collections
import statistics
import os
import json
import re
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Characters illegal in Excel/XML cell values (openpyxl raises IllegalCharacterError)
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def _clean(value):
    """Strip illegal XML characters from strings before writing to Excel."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub("", value)
    return value

try:
    import boto3
    import pyreadr
    _S3_AVAILABLE = True
except ImportError:
    _S3_AVAILABLE = False

# ── paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
SAVE_DIR    = os.path.join(SCRIPT_DIR, "..", "..", "..",
    "Laura Pyle's files - Biostatistics Core Shared Drive",
    "Data Harmonization", "Data Clean")
DATA_FILE   = os.path.join(
    SCRIPT_DIR, "..", "..", "..",
    "Laura Pyle's files - Biostatistics Core Shared Drive",
    "Data Harmonization", "Data Clean",
    "soma_olink_harmonized_dataset.csv",
)
OUTPUT_FILE    = os.path.join(SAVE_DIR, "Data_Inventory.xlsx")
DATA_DICT_FILE = os.path.join(
    SCRIPT_DIR, "..", "..", "..",
    "Laura Pyle's files - Biostatistics Core Shared Drive",
    "Data Harmonization",
    "data_dictionary_master.csv",
)

# ── S3 / PB90 scRNA constants ─────────────────────────────────────────────────
KEYS_FILE   = "/Users/shivaniramesh/Desktop/keys.json"
S3_ENDPOINT = "https://s3.kopah.uw.edu"
S3_BUCKET   = "scrna"
S3_PB90_KEY = "data_clean/pb90_meta.rds"

# Maps PB90 cohort labels → harmonized dataset study names
COHORT_MAP = {
    "ATTEMPT":          "ATTEMPT",
    "CROCODILE":        "CROCODILE",
    "IMPROVE":          "IMPROVE",
    "PANDA":            "PANDA",
    "RENAL HEIR":       "RENAL-HEIR",
    "RENAL HEIRITAGE":  "RENAL-HEIRitage",
}

# Study-specific visit allowlists.  Rows whose visit value is not in the set
# are skipped entirely (subjects, proc_n, Table 1, omics, proc_cols).
# Studies not listed here keep all visits.
STUDY_VISIT_FILTER = {
    "ATTEMPT": {"screening", "baseline", "4_weeks_post", "4_months_post"},
}


# form_name values in data_dictionary_master.csv that represent
# demographics, consent, baseline admin data — excluded from the
# per-study dictionary panels.  Extend as needed.
EXCLUDED_DICT_FORMS = {
    "demographics", "consent", "informed_consent",
    "inclusion_exclusion", "eligibility", "screening",
    "enrollment", "study_status", "participant_status",
    "medical_history", "past_medical_history",
    "medications", "concomitant_medications", "adverse_events",
    "withdrawal", "termination", "contact_information",
    "race_ethnicity", "baseline",
    # additional admin/non-procedure forms

}
# Variable names that are always excluded from procedure dictionary panels,
# regardless of form_name.  Covers demographic / operational / admin variables
# whose form_name is not captured by EXCLUDED_DICT_FORMS, or that are
# undocumented and would otherwise slip in via the third-pass bare-name logic.
EXCLUDED_VARS = {
    # demographic / admin
    "diabetes_duration", "group_risk", "participation_status", "race_ethnicity",
    "birthweight",
    # lab / clinical values that bleed into non-lab procedure rows
    "cystatin_c_s", "eGFR_fas_cr_cysc",
    # medication flag columns (yes/no fields, not procedure measurements)
    "fibrates", "mra", "phentermine", "topiramate",
}

# Set True to print per-study form→procedure mapping and proc_groups to stdout.
DEBUG_DICT = False

# Word-level synonyms used when matching REDCap form_names to procedure names.
# Key = word that may appear in a form_name; value = the procedure name word
# it is equivalent to.  Only single-word mappings needed.
_WORD_SYNONYMS = {
    "meds": "medications",
    "med":  "medications",
    "boldasl": "bold",   # study_visit_boldasl_mri → bold_mri
}


# ── Table 1 variable definitions ─────────────────────────────────────────────
# (col_name, display_label, stat)  stat in {'mean', 'median', 'pct_female', 'category'}
TABLE1_VARS = [
    ("group",       "Group",                          "category"),
    ("age",         "Age (years)",                    "mean"),
    ("sex",         "Sex (% Female)",                 "pct_female"),
    ("bmi",         "BMI (kg/m²)",                    "mean"),
    ("acr_u",       "UACR (mg/g)",                    "median"),
    ("eGFR_CKD_epi","eGFR CKD-EPI (mL/min/1.73 m²)", "mean"),
    ("eGFR_fas_cr", "eGFR FAS-Cr (mL/min/1.73 m²)",  "mean"),
]

# Columns excluded from the ≥3 non-missing check
KEY_COLS = {
    "record_id","attempt_id","casper_id","coffee_id","croc_id",
    "improve_id","penguin_id","rh_id","rh2_id","panther_id",
    "panda_id","rpc2_id","swht_id","ultra_id","co_enroll_id","mrn",
    "date","screen_date","date_of_screen","study","dob","diabetes_dx_date",
    "sex","race","ethnicity","visit","procedure","group","age",
}

MISSING_VALUES = {"", "NA", "na", "NaN", "nan", "N/A", "n/a", "NULL", "null"}

# ── colour palette ────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F7"
WHITE        = "FFFFFF"
GREY         = "F2F2F2"
LIGHT_GREEN  = "E2EFDA"

# ── style helpers ─────────────────────────────────────────────────────────────

def _side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

def _border(style="thin", color="BFBFBF"):
    s = _side(style, color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=11, bold=False, color="000000", italic=False):
    return Font(name="Calibri", size=size, bold=bold,
                color=color, italic=italic)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_cell(cell, value=None, *, bg=WHITE, fg="000000", size=11,
               bold=False, italic=False, h_align="left", v_align="center",
               wrap=True, border_style="thin", border_color="BFBFBF"):
    if value is not None:
        cell.value = _clean(value)
    cell.font      = _font(size=size, bold=bold, color=fg, italic=italic)
    cell.fill      = _fill(bg)
    cell.border    = _border(border_style, border_color)
    cell.alignment = _align(h_align, v_align, wrap)


def title_row(ws, row, col_start, col_end, text, *, bg=DARK_BLUE,
              fg=WHITE, size=13, height=28):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    style_cell(ws.cell(row=row, column=col_start), text,
               bg=bg, fg=fg, size=size, bold=True, h_align="center")
    ws.row_dimensions[row].height = height


def section_row(ws, row, col_start, col_end, text, *, bg=MID_BLUE, height=20):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    style_cell(ws.cell(row=row, column=col_start), text,
               bg=bg, fg=WHITE, size=11, bold=True)
    ws.row_dimensions[row].height = height


def header_row(ws, row, labels, *, col_start=1, bg=MID_BLUE,
               fg=WHITE, height=20):
    for i, lbl in enumerate(labels):
        style_cell(ws.cell(row=row, column=col_start + i), lbl,
                   bg=bg, fg=fg, size=10, bold=True, h_align="center")
    ws.row_dimensions[row].height = height


def blank_row(ws, row, height=8):
    ws.row_dimensions[row].height = height


# ── statistics helpers ────────────────────────────────────────────────────────

def _safe_float(v):
    """Return float or None."""
    v = str(v).strip().strip('"')
    if v in MISSING_VALUES:
        return None
    try:
        return float(v)
    except ValueError:
        return None


def compute_stats(values, stat):
    """Return (n, formatted_value) for a list of raw string values."""
    if stat == "pct_female":
        clean = [v.strip().strip('"') for v in values
                 if v.strip().strip('"') not in MISSING_VALUES]
        n = len(clean)
        if n == 0:
            return 0, "—"
        pct = sum(1 for v in clean if v.lower() == "female") / n * 100
        return n, f"{pct:.0f}%"
    elif stat == "category":
        return len(values), "—"
    else:
        nums = [_safe_float(v) for v in values]
        nums = [x for x in nums if x is not None]
        n = len(nums)
        if n == 0:
            return 0, "—"
        if stat == "mean":
            return n, f"{statistics.mean(nums):.1f}"
        elif stat == "median":
            return n, f"{statistics.median(nums):.1f}"
    return 0, "—"


# ── load omics column names from data dictionary ──────────────────────────────

def load_omics_cols(dict_file, all_cols):
    """
    Read data_dictionary_master.csv and return a set of metabolomics column
    names that are also present in all_cols (the dataset's fieldnames).
    """
    metab_forms = {
        "az_urine_metabolites", "metabolomics", "metabolomics_aq",
        "metabolomics_blood_raw", "metabolomics_tissue",
    }
    metab_vars = set()
    try:
        with open(dict_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                form = row.get("form_name", "").strip().lower()
                var  = row.get("variable_name", "").strip()
                if form in metab_forms and var:
                    metab_vars.add(var)
    except FileNotFoundError:
        print(f"  WARNING: Data dictionary not found — metabolomics cols skipped.")
        return set()

    metab_cols = metab_vars & set(all_cols)
    print(f"  Data dictionary: {len(metab_vars)} metabolomics vars identified, "
          f"{len(metab_cols)} present in dataset.")
    return metab_cols


# ── load full data dictionary ─────────────────────────────────────────────────

def load_data_dictionary(dict_file):
    """
    Read data_dictionary_master.csv.
    Returns: { variable_name: {'label': str, 'units': str, 'form_name': str} }
    """
    data_dict = {}
    try:
        with open(dict_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                var = row.get("variable_name", "").strip()
                if var:
                    data_dict[var] = {
                        "label":     row.get("label",     "").strip(),
                        "units":     row.get("units",     "").strip(),
                        "form_name": row.get("form_name", "").strip(),
                    }
    except FileNotFoundError:
        print(f"  WARNING: Data dictionary not found — dict section will be skipped.")
    print(f"  Data dictionary loaded: {len(data_dict):,} variables.")
    return data_dict


# ── load PB90 scRNA metadata from Kopah S3 ───────────────────────────────────

def load_scrna_metadata():
    """
    Download pb90_meta.rds from Kopah S3 and return:

        {
          study_name: {
            'n_cells':    int,
            'n_subjects': int,
            'by_group': {
                group_label: {
                    'n_cells': int,
                    'cell_types': { celltype: n_cells, ... },
                }
            },
            'cell_types': { celltype: n_cells, ... },   # overall counts
          }
        }

    Keys use the harmonised study names (via COHORT_MAP).
    Returns empty dict on any error.
    """
    if not _S3_AVAILABLE:
        print("  WARNING: boto3/pyreadr not installed — skipping PB90 download.")
        print("           Install with: pip install boto3 pyreadr")
        return {}

    try:
        keys = json.load(open(KEYS_FILE))
        session = boto3.Session(
            aws_access_key_id=keys["MY_ACCESS_KEY"],
            aws_secret_access_key=keys["MY_SECRET_KEY"],
        )
        s3 = session.client("s3", endpoint_url=S3_ENDPOINT)

        tmp = tempfile.NamedTemporaryFile(suffix=".rds", delete=False)
        tmp.close()
        print("  Downloading pb90_meta.rds from Kopah S3 …")
        s3.download_file(S3_BUCKET, S3_PB90_KEY, tmp.name)

        result = pyreadr.read_r(tmp.name)
        os.unlink(tmp.name)
        df = result[None] if None in result else next(iter(result.values()))
        print(f"  PB90 loaded: {len(df):,} cells")

    except Exception as exc:
        print(f"  WARNING: could not load PB90 — {exc}")
        return {}

    scrna = {}
    for cohort_val, cohort_df in df.groupby("cohort"):
        study = COHORT_MAP.get(str(cohort_val))
        if study is None:
            continue

        entry = {
            "n_cells":    len(cohort_df),
            "n_subjects": cohort_df["record_id"].nunique(),
            "by_group":   {},
            "cell_types": cohort_df["celltype_rpca"].value_counts().to_dict(),
        }
        for grp_val, grp_df in cohort_df.groupby("group"):
            entry["by_group"][str(grp_val)] = {
                "n_cells":    len(grp_df),
                "n_subjects": grp_df["record_id"].nunique(),
                "cell_types": grp_df["celltype_rpca"].value_counts().to_dict(),
            }
        scrna[study] = entry

    return scrna


# ── read dataset ──────────────────────────────────────────────────────────────

def load_study_metadata(data_file):
    """
    Single-pass CSV read.  Returns:

    study_meta : dict  study_name -> {
        'subjects'  : set of record_ids,
        'visits'    : sorted list,
        'groups'    : sorted list,
        'procedures': sorted list,
        'proc_n'    : dict  procedure -> # unique record_ids with data values,
        'table1'    : dict  group_label -> { col: [raw_values, ...] }
                      (one value per record_id, first non-missing wins)
    }
    soma_cols, olink_cols, clinical_cols : lists of column names
    """
    study_meta = collections.defaultdict(lambda: {
        "subjects": set(), "visits": set(),
        "groups": set(), "procedures": set(),
        "proc_n": {},
        "table1": collections.defaultdict(
            lambda: collections.defaultdict(list)),
        "omics": {
            "soma": set(), "olink_p": set(), "olink_u": set(),
            "metab": set(), "lipid": set(),
        },
        "study_cols": set(),  # clinical col names with ≥1 non-missing value
    })

    soma_cols        = []
    olink_cols       = []
    olink_p_cols     = []   # OID*_p  — plasma
    olink_u_cols     = []   # OID*_u  — urine
    clinical_cols    = []

    # For proc_n: unique record_ids per (study, proc)
    qualified = collections.defaultdict(set)   # (study, proc) -> set of record_ids
    # For instructions matrix: unique record_ids per (proc, group) across all studies
    proc_group_qualified = collections.defaultdict(set)  # (proc, group) -> set of record_ids

    # For data dictionary: which clinical cols have ≥1 non-missing value per study
    study_col_tracker = collections.defaultdict(set)
    # Per-procedure: which clinical cols have ≥1 non-missing value for each procedure
    proc_col_tracker  = collections.defaultdict(set)  # (study, proc) -> set of col names
    # Per-variable participant counts: (study, proc, col) -> set of record_ids with non-missing value
    var_n_tracker     = collections.defaultdict(set)  # (study, proc, col) -> set of record_ids
    # No-procedure row tracking: cols with data where procedure is blank
    no_proc_col_tracker  = collections.defaultdict(set)  # study -> set of col names
    no_proc_var_n_tracker = collections.defaultdict(set) # (study, col) -> set of record_ids


    # For Table 1: track first non-missing value per (study, record_id, col)
    t1_cols   = [col for col, _, _ in TABLE1_VARS if col != "group"]
    seen_t1   = collections.defaultdict(dict)   # (study, rid) -> {col: value}
    rid_group = {}                              # (study, rid) -> first group seen

    t1_col_set = set(t1_cols)

    with open(data_file, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [c.strip('"') for c in reader.fieldnames]

        for col in reader.fieldnames:
            if col.startswith("seq."):
                soma_cols.append(col)
            elif col.startswith("OID"):
                olink_cols.append(col)
                if col.endswith("_p"):
                    olink_p_cols.append(col)
                elif col.endswith("_u"):
                    olink_u_cols.append(col)
            elif col not in KEY_COLS:
                clinical_cols.append(col)

        metab_cols = load_omics_cols(DATA_DICT_FILE, reader.fieldnames)
        print(f"  SomaScan cols: {len(soma_cols):,}  |  "
              f"Olink plasma: {len(olink_p_cols):,}  |  "
              f"Olink urine: {len(olink_u_cols):,}  |  "
              f"Metabolomics: {len(metab_cols):,}")


        for row in reader:
            study = row.get("study", "").strip().strip('"')
            if not study:
                continue

            rid   = row.get("record_id", "").strip().strip('"')
            visit = row.get("visit",     "").strip().strip('"')
            group = row.get("group",     "").strip().strip('"')
            proc  = row.get("procedure", "").strip().strip('"')

            # Apply study-specific visit allowlist
            visit_filter = STUDY_VISIT_FILTER.get(study)
            if visit_filter is not None and visit not in visit_filter:
                continue

            m = study_meta[study]
            # Always collect visits/groups/procedures regardless of enrollment
            if visit: m["visits"].add(visit)
            if group: m["groups"].add(group)
            if proc:  m["procedures"].add(proc)

            # Any row with a record_id is counted — no study-specific ID gate.
            if not rid:
                continue

            m["subjects"].add(rid)

            # ── per-procedure clinical col tracking (for dictionary) ───────────
            if proc:
                found_p = proc_col_tracker[(study, proc)]
                for c in clinical_cols:
                    val = row.get(c, "").strip().strip('"')
                    if val not in MISSING_VALUES:
                        found_p.add(c)
                        if rid:
                            var_n_tracker[(study, proc, c)].add(rid)
            else:
                # Track clinical cols with data in rows that have no procedure
                found_np = no_proc_col_tracker[study]
                for c in clinical_cols:
                    val = row.get(c, "").strip().strip('"')
                    if val not in MISSING_VALUES:
                        found_np.add(c)
                        if rid:
                            no_proc_var_n_tracker[(study, c)].add(rid)


            # ── proc_n: participants with ≥1 non-missing clinical value ────
            if proc:
                has_clinical = any(
                    row.get(c,"").strip().strip('"') not in MISSING_VALUES
                    for c in clinical_cols
                )
                if has_clinical:
                    qualified[(study, proc)].add(rid)
                    if group:
                        proc_group_qualified[(proc, group)].add(rid)

            # ── Table 1: accumulate one value per (study, rid) ────────────
            if group:
                key = (study, rid)
                if key not in rid_group:
                    rid_group[key] = group          # store first group seen
                seen = seen_t1[key]
                for col in t1_col_set:
                    if col not in seen:
                        val = row.get(col, "").strip().strip('"')
                        if val not in MISSING_VALUES:
                            seen[col] = val

            # ── omics data tracking ────────────────────────────────────────
            om = m["omics"]
            # Lipidomics: identified by procedure label
            if proc and proc.lower() == "lipidomics":
                om["lipid"].add(rid)
            # Column-based omics: short-circuit once confirmed for this rid
            if rid not in om["soma"]:
                for c in soma_cols:
                    if row.get(c, "").strip().strip('"') not in MISSING_VALUES:
                        om["soma"].add(rid)
                        break
            if rid not in om["olink_p"]:
                for c in olink_p_cols:
                    if row.get(c, "").strip().strip('"') not in MISSING_VALUES:
                        om["olink_p"].add(rid)
                        break
            if rid not in om["olink_u"]:
                for c in olink_u_cols:
                    if row.get(c, "").strip().strip('"') not in MISSING_VALUES:
                        om["olink_u"].add(rid)
                        break
            if rid not in om["metab"] and metab_cols:
                for c in metab_cols:
                    if row.get(c, "").strip().strip('"') not in MISSING_VALUES:
                        om["metab"].add(rid)
                        break

            # ── clinical column presence: study-level ─────────────────────
            found_s = study_col_tracker[study]
            for c in clinical_cols:
                val = row.get(c, "").strip().strip('"')
                if val not in MISSING_VALUES:
                    found_s.add(c)

    # attach proc_n
    for (study, proc), rid_set in qualified.items():
        study_meta[study]["proc_n"][proc] = len(rid_set)

    # attach study_cols
    for study, col_set in study_col_tracker.items():
        study_meta[study]["study_cols"] = col_set

    # attach proc_cols
    for (study, proc), col_set in proc_col_tracker.items():
        if "proc_cols" not in study_meta[study]:
            study_meta[study]["proc_cols"] = {}
        study_meta[study]["proc_cols"][proc] = col_set

    # attach var_n: (proc, col) -> participant count
    for (study, proc, col), rid_set in var_n_tracker.items():
        if "var_n" not in study_meta[study]:
            study_meta[study]["var_n"] = {}
        study_meta[study]["var_n"][(proc, col)] = len(rid_set)

    # attach no_proc_cols and no-proc var_n counts
    for study, col_set in no_proc_col_tracker.items():
        study_meta[study]["no_proc_cols"] = col_set
    for (study, col), rid_set in no_proc_var_n_tracker.items():
        if "var_n" not in study_meta[study]:
            study_meta[study]["var_n"] = {}
        study_meta[study]["var_n"][("", col)] = len(rid_set)

    # build Table 1 structures  ->  group -> col -> [values]
    for (study, rid), vals in seen_t1.items():
        group = rid_group.get((study, rid), "Unknown")
        t1    = study_meta[study]["table1"]
        for col in t1_col_set:
            val = vals.get(col, "")
            t1[group][col].append(val)

    # sort / finalise
    for m in study_meta.values():
        m["visits"]     = sorted(m["visits"])
        m["groups"]     = sorted(m["groups"])
        m["procedures"] = sorted(m["procedures"])
        m["proc_n"]     = dict(m["proc_n"])
        m["table1"]     = dict(m["table1"])
        # Convert omics sets → counts; compute union for "any omics"
        om = m["omics"]
        any_omics = (om["soma"] | om["olink_p"] | om["olink_u"]
                     | om["metab"] | om["lipid"])
        m["omics"] = {k: len(v) for k, v in om.items()}
        m["omics"]["any"] = len(any_omics)

    # build proc_group_n: proc -> group -> count
    proc_group_n = collections.defaultdict(dict)
    for (proc, group), rids in proc_group_qualified.items():
        proc_group_n[proc][group] = len(rids)

    return (dict(sorted(study_meta.items())),
            soma_cols, olink_cols, clinical_cols,
            dict(proc_group_n))


# ── build workbook ────────────────────────────────────────────────────────────

def build_workbook(study_meta, soma_cols, olink_cols, clinical_cols,
                   scrna_meta=None, data_dict=None, proc_group_n=None):
    if scrna_meta is None:
        scrna_meta = {}
    if data_dict is None:
        data_dict = {}
    if proc_group_n is None:
        proc_group_n = {}
    wb = Workbook()
    wb.remove(wb.active)
    _add_instructions_sheet(wb, study_meta, proc_group_n)
    for study_name, meta in study_meta.items():
        _add_study_sheet(wb, study_name, meta,
                         soma_cols, olink_cols, clinical_cols,
                         scrna_meta.get(study_name),
                         data_dict=data_dict)
    return wb


# ── instructions sheet ────────────────────────────────────────────────────────

def _add_instructions_sheet(wb, study_meta, proc_group_n=None):
    from openpyxl.utils import get_column_letter
    if proc_group_n is None:
        proc_group_n = {}

    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    # ── derive column layout for wide procedure matrix ──────────────────────
    all_studies = list(study_meta.keys())          # already sorted
    all_groups  = sorted({g for m in study_meta.values() for g in m["groups"]})
    all_procs   = sorted({p for m in study_meta.values()
                           for p in m.get("proc_n", {}).keys()})

    # col B = procedure name, then one col per study, then one per group
    n_studies   = len(all_studies)
    n_groups    = len(all_groups)
    proc_col    = 2                          # column B
    study_start = 3                          # columns C …
    group_start = study_start + n_studies    # … then groups
    tbl_end     = group_start + n_groups - 1 # last column of the table

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions[get_column_letter(proc_col)].width = 28
    for i in range(n_studies + n_groups):
        ws.column_dimensions[get_column_letter(study_start + i)].width = 13

    r = 1
    title_row(ws, r, proc_col, tbl_end,
              "SomaScan / Olink Harmonized Dataset — Study Reference Summary",
              size=15, height=36); r += 1
    title_row(ws, r, proc_col, tbl_end,
              "Biostatistics Core – University of Washington",
              bg=MID_BLUE, size=11, height=22); r += 1
    blank_row(ws, r); r += 1

    section_row(ws, r, proc_col, tbl_end, "PURPOSE"); r += 1
    ws.merge_cells(start_row=r, start_column=proc_col,
                   end_row=r, end_column=tbl_end)
    ws.row_dimensions[r].height = 130
    style_cell(
        ws.cell(row=r, column=proc_col),
        (
            "This workbook is a reference summary of the harmonised "
            "SomaScan / Olink dataset (soma_olink_harmonized_dataset.csv), "
            "intended to support Google Form–based data subset requests.\n\n"
            "Master dataset:  ~15,900 rows  |  28,673 columns\n"
            "  • ~25,354 SomaScan analytes  (seq.*)\n"
            "  • ~1,656 Olink analytes  (OID*_u)\n"
            "  • ~1,634 clinical / lab / imaging variables\n\n"
            "Each study tab contains:\n"
            "  • Study overview  (N, visits, groups, procedures)\n"
            "  • Procedure participation  (# participants with values)\n"
            "  • Table 1 characteristics  (age, sex, BMI, UACR, eGFR) stratified by group\n"
            "  • PB90 scRNA inventory  (cell counts by cell type and group)"
        ),
        bg=GREY, size=11, v_align="top",
        border_style="medium", border_color=MID_BLUE,
    ); r += 1
    blank_row(ws, r); r += 1

    section_row(ws, r, proc_col, tbl_end, "COLOUR KEY"); r += 1
    for bg, fg, txt in [
        (LIGHT_GREEN, "000000", "Green  =  reference data (auto-populated from master dataset)"),
        (LIGHT_BLUE,  "000000", "Blue   =  alternating data rows"),
    ]:
        ws.row_dimensions[r].height = 20
        ws.merge_cells(start_row=r, start_column=proc_col,
                       end_row=r, end_column=tbl_end)
        style_cell(ws.cell(row=r, column=proc_col), txt, bg=bg, fg=fg, size=10)
        r += 1

    blank_row(ws, r); r += 1

    # ── Study sheets summary table (narrow — just name + N) ────────────────
    section_row(ws, r, proc_col, tbl_end, "STUDY SHEETS"); r += 1
    header_row(ws, r, ["Study", "# Participants"], col_start=proc_col); r += 1
    total_participants = 0
    for i, (sname, meta) in enumerate(study_meta.items()):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20
        style_cell(ws.cell(row=r, column=proc_col), sname, bg=bg, bold=True)
        n = len(meta["subjects"])
        total_participants += n
        style_cell(ws.cell(row=r, column=study_start), n,
                   bg=bg, h_align="center")
        r += 1
    ws.row_dimensions[r].height = 20
    style_cell(ws.cell(row=r, column=proc_col), "Total",
               bg=DARK_BLUE, fg=WHITE, bold=True, size=11)
    style_cell(ws.cell(row=r, column=study_start), total_participants,
               bg=DARK_BLUE, fg=WHITE, h_align="center", bold=True, size=11)
    r += 1

    blank_row(ws, r); r += 1

    # ── All-studies procedure matrix ────────────────────────────────────────
    # Rows = procedures; cols = one per study + one per group
    section_row(ws, r, proc_col, tbl_end,
                "ALL PROCEDURES  —  # participants with ≥1 non-missing clinical value"); r += 1

    # Header: Procedure | Study1 | Study2 | … | [Group1 | Group2 | …]
    hdr_labels = ["Procedure"] + all_studies + all_groups
    for j, lbl in enumerate(hdr_labels):
        c = proc_col + j
        # Shade study headers slightly different from group headers
        bg = MID_BLUE if j == 0 or j <= n_studies else DARK_BLUE
        style_cell(ws.cell(row=r, column=c), lbl,
                   bg=bg, fg=WHITE, size=9, bold=True,
                   h_align="center", wrap=True)
    ws.row_dimensions[r].height = 36
    r += 1

    for i, proc in enumerate(all_procs):
        bg = LIGHT_GREEN if i % 2 == 0 else WHITE

        style_cell(ws.cell(row=r, column=proc_col), proc,
                   bg=bg, fg=DARK_BLUE, size=10)

        # one cell per study
        for j, study in enumerate(all_studies):
            n = study_meta[study].get("proc_n", {}).get(proc, "")
            style_cell(ws.cell(row=r, column=study_start + j),
                       n if n != "" else "—",
                       bg=bg, h_align="center", size=10)

        # one cell per group
        for j, grp in enumerate(all_groups):
            n = proc_group_n.get(proc, {}).get(grp, "")
            style_cell(ws.cell(row=r, column=group_start + j),
                       n if n != "" else "—",
                       bg=bg, h_align="center", size=10)
        r += 1


# ── per-study sheet ───────────────────────────────────────────────────────────

def _add_study_sheet(wb, study_name, meta,
                     soma_cols, olink_cols, clinical_cols,
                     scrna=None, data_dict=None):
    from openpyxl.utils import get_column_letter

    safe_name = study_name.replace("/", "-")[:31]
    ws = wb.create_sheet(safe_name)
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 3, "B": 32, "C": 18, "D": 18,
                  "E": 18, "F": 18, "G": 3}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── pre-compute dictionary column positions ─────────────────────────────
    # Dictionary sits to the right of the widest left-side section.
    _n_grps_pre      = len(meta["groups"])
    _t1_end          = 2 + 1 + _n_grps_pre * 2 + 2
    _n_scrna_pre     = len(scrna["by_group"]) if scrna else 0
    _pb90_end        = max(3 + _n_scrna_pre, 6)
    _left_end        = max(_t1_end, _pb90_end)
    DICT_GAP_COL     = _left_end + 1   # narrow spacer column
    DICT_VAR_COL     = DICT_GAP_COL + 1
    DICT_LABEL_COL   = DICT_VAR_COL + 1
    DICT_UNITS_COL   = DICT_LABEL_COL + 1
    DICT_FORM_COL    = DICT_UNITS_COL + 1
    DICT_N_COL       = DICT_FORM_COL + 1

    ws.column_dimensions[get_column_letter(DICT_GAP_COL)].width   = 3
    ws.column_dimensions[get_column_letter(DICT_VAR_COL)].width   = 30
    ws.column_dimensions[get_column_letter(DICT_LABEL_COL)].width = 55
    ws.column_dimensions[get_column_letter(DICT_UNITS_COL)].width = 14
    ws.column_dimensions[get_column_letter(DICT_FORM_COL)].width  = 22
    ws.column_dimensions[get_column_letter(DICT_N_COL)].width     = 14

    r = 1

    # ── pre-compute procedure-grouped dictionary sections ──────────────────
    # Uses EMPIRICAL per-procedure column tracking (proc_col_tracker from the
    # data-loading phase): for each procedure, we directly know which clinical
    # columns had ≥1 non-missing value in rows where that procedure was recorded.
    # We then look each column up in the data dictionary, skipping columns
    # whose form_name is in EXCLUDED_DICT_FORMS *unless* the form_name exactly
    # matches the procedure name (self-match — e.g. "physical_exam" form under
    # the "physical_exam" procedure).
    proc_groups    = {}   # proc_name -> [(col_name, info), ...]
    form_dict_rows = {}   # proc_name -> first row number of its section
    no_proc_vars   = []   # [(col_name, info), ...] for no-procedure rows

    if data_dict:
        study_procs     = [p for p in meta["procedures"] if p]
        study_proc_cols = meta.get("proc_cols", {})

        # ── Step 1: build per-procedure candidate variable lists ──────────────
        # raw_proc_groups: exclusion-filtered (EXCLUDED_DICT_FORMS applied)
        # full_proc_groups: all data_dict vars, no exclusion filter (fallback)
        raw_proc_groups  = {}
        full_proc_groups = {}

        for proc in sorted(study_procs):
            proc_col_set = study_proc_cols.get(proc, set())
            proc_lower   = proc.lower()
            _filtered, _full = [], []
            for col_name in sorted(proc_col_set):
                if col_name in EXCLUDED_VARS:
                    continue
                if col_name not in data_dict:
                    continue
                info     = data_dict[col_name]
                fn_lower = (info.get("form_name", "") or "").lower()
                _full.append((col_name, info))
                if fn_lower not in EXCLUDED_DICT_FORMS or fn_lower == proc_lower:
                    _filtered.append((col_name, info))
            raw_proc_groups[proc]  = _filtered
            full_proc_groups[proc] = _full

        # ── Step 2: deduplicate — assign each variable to exactly one proc ────
        # Priority: (1) exact form_name match; (2) highest var_n among procs
        # where the variable is not cross-cutting (≤ half of all procedures).
        var_to_proc = {}
        n_procs = len(study_procs)
        _cross_thresh = max(1, n_procs // 2)
        var_n_lookup  = meta.get("var_n", {})

        # Count how many procedures each documented variable appears in.
        # Variables present in more than half the procedures are cross-cutting
        # (filled in on every row regardless of procedure) and should not be
        # forced onto any one procedure in the second pass.
        var_proc_count = collections.Counter()
        for proc in study_procs:
            for col_name in study_proc_cols.get(proc, set()):
                if col_name in data_dict and col_name not in EXCLUDED_VARS:
                    var_proc_count[col_name] += 1

        # First pass: exact form_name → proc name matches take priority
        for proc in sorted(study_procs):
            proc_lower = proc.lower()
            for col_name, info in raw_proc_groups.get(proc, []):
                if (info.get("form_name", "") or "").lower() == proc_lower:
                    var_to_proc[col_name] = proc

        # Second pass: for each remaining variable, collect candidate procedures
        # (those where var_proc_count ≤ threshold) and assign to the one with
        # the highest unique-participant count.  Alphabetical name breaks ties.
        _var_candidates = collections.defaultdict(list)
        for proc in sorted(study_procs):
            for col_name, info in raw_proc_groups.get(proc, []):
                if col_name not in var_to_proc:
                    if var_proc_count.get(col_name, 1) <= _cross_thresh:
                        n = var_n_lookup.get((proc, col_name), 0)
                        _var_candidates[col_name].append((n, proc))
        for col_name, candidates in _var_candidates.items():
            if col_name not in var_to_proc and candidates:
                # highest var_n wins; alphabetical proc name as tiebreaker
                candidates.sort(key=lambda x: (-x[0], x[1]))
                var_to_proc[col_name] = candidates[0][1]

        # Third pass: data-tracked vars absent from data_dict (bare names).
        # Only include vars tracked by exactly one procedure in this study —
        # cross-procedure vars (race___, ethnicity___, uuid, etc.) are filled in
        # on every row and should not be pulled into a procedure's dictionary.
        undoc_proc_count = collections.Counter()
        for proc in study_procs:
            for col_name in study_proc_cols.get(proc, set()):
                if (col_name not in data_dict
                        and col_name not in var_to_proc
                        and col_name not in EXCLUDED_VARS):
                    undoc_proc_count[col_name] += 1
        for proc in sorted(study_procs):
            for col_name in sorted(study_proc_cols.get(proc, set())):
                if (col_name not in data_dict
                        and col_name not in var_to_proc
                        and col_name not in EXCLUDED_VARS
                        and undoc_proc_count[col_name] == 1):
                    var_to_proc[col_name] = proc

        # Build proc_groups from assignment map
        proc_groups = {proc: [] for proc in sorted(study_procs)}
        for col_name, best_proc in var_to_proc.items():
            proc_groups[best_proc].append((col_name, data_dict.get(col_name, {})))
        for proc in proc_groups:
            proc_groups[proc].sort(key=lambda x: x[0])

        # ── Step 3: QC — ensure no procedure section is empty ─────────────────
        for proc in sorted(study_procs):
            if proc_groups[proc]:
                continue
            # Fallback 1: include full (non-exclusion-filtered) unassigned vars,
            # but still skip cross-cutting vars (appear in > half of procedures).
            for col_name, info in full_proc_groups.get(proc, []):
                if (col_name not in var_to_proc
                        and col_name not in EXCLUDED_VARS
                        and var_proc_count.get(col_name, 1) <= _cross_thresh):
                    proc_groups[proc].append((col_name, info))
                    var_to_proc[col_name] = proc
            proc_groups[proc].sort(key=lambda x: x[0])
            # Fallback 2: include proc_col_set vars not in data_dict (bare names),
            # again skipping cross-cutting undocumented vars.
            if not proc_groups[proc]:
                proc_col_set = study_proc_cols.get(proc, set())
                unassigned = sorted(c for c in proc_col_set
                                    if c not in var_to_proc
                                    and c not in EXCLUDED_VARS
                                    and undoc_proc_count.get(c, 1) <= _cross_thresh)
                proc_groups[proc] = [(c, {}) for c in unassigned]
                for c in unassigned:
                    var_to_proc[c] = proc

        if DEBUG_DICT:
            print(f"  [{study_name}] proc_groups (deduplicated): "
                  + ", ".join(f"{p}({len(v)})" for p, v in proc_groups.items()))
            
        # ── Build no-procedure variables section ───────────────────────────────
        # Includes clinical cols with non-missing data in rows where procedure=""
        # that are NOT in EXCLUDED_DICT_FORMS.  Undocumented cols are shown as
        # bare names.  Admin/key cols are already absent from no_proc_cols since
        # they were excluded from clinical_cols at load time.
        no_proc_cols_set = meta.get("no_proc_cols", set())
        already_assigned = set(var_to_proc.keys())
        for col_name in sorted(no_proc_cols_set):
            info = data_dict.get(col_name)
            if info is not None:
                fn_lower = (info.get("form_name", "") or "").lower()
                if fn_lower in EXCLUDED_DICT_FORMS:
                    continue
                no_proc_vars.append((col_name, info))
            else:
                # Undocumented variable — include as bare name
                no_proc_vars.append((col_name, {}))

        # Pre-compute dr row positions for hyperlinks.
        # Empty sections: sub-header + note row + spacer = 3 rows.
        # Non-empty: sub-header + col-header + N var rows + spacer = N+3.
        _dr = 3   # title=row1, blank=row2, first section starts at row3
        for proc_name, pvars in proc_groups.items():
            form_dict_rows[proc_name] = _dr
            _dr += 1            # sub-header row
            if pvars:
                _dr += 1            # column-header row
                _dr += len(pvars)   # one row per variable
            else:
                _dr += 1            # note row
            _dr += 1            # blank spacer row

    # Direct procedure → row lookup
    def _proc_form_row(proc):
        row = form_dict_rows.get(proc)
        return (proc, row) if row is not None else (None, None)

    # ── title ──────────────────────────────────────────────────────────────
    title_row(ws, r, 2, 6,
              f"Study Summary  —  {study_name}",
              size=14, height=30); r += 1
    blank_row(ws, r); r += 1

    # ── study overview ─────────────────────────────────────────────────────
    section_row(ws, r, 2, 6, "STUDY OVERVIEW"); r += 1

    summary_items = [
        ("Study",              study_name),
        ("Total Participants", str(len(meta["subjects"]))),
        ("Available Visits",   ", ".join(meta["visits"]) if meta["visits"] else "—"),
        ("Participant Groups",  ", ".join(meta["groups"])),
        ("Procedures",          ", ".join(meta["procedures"])),
    ]
    for i, (k, v) in enumerate(summary_items):
        bg = LIGHT_GREEN if i % 2 == 0 else WHITE
        style_cell(ws.cell(row=r, column=2), k, bg=bg, fg=DARK_BLUE, bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        style_cell(ws.cell(row=r, column=3), v, bg=bg)
        r += 1

    blank_row(ws, r, height=8); r += 1

    # ── procedure participation table ──────────────────────────────────────
    section_row(ws, r, 2, 6,
                "PARTICIPANTS WITH DATA PER PROCEDURE  "
                "(unique record_id with values in that row)"); r += 1

    header_row(ws, r, ["Procedure", "# Participants with Data"],
               col_start=2, height=20)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    r += 1

    proc_n = meta.get("proc_n", {})
    for i, proc in enumerate(meta["procedures"]):
        bg = LIGHT_GREEN if i % 2 == 0 else WHITE
        proc_cell = ws.cell(row=r, column=2)
        style_cell(proc_cell, proc, bg=bg, fg=DARK_BLUE)
        _matched_fn, _matched_row = _proc_form_row(proc)
        if _matched_row is not None:
            proc_cell.hyperlink = (
                f"#'{ws.title}'!"
                f"{get_column_letter(DICT_VAR_COL)}{_matched_row}"
            )
            proc_cell.font = Font(name="Calibri", size=11, bold=False,
                                  color="1155CC", underline="single")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        style_cell(ws.cell(row=r, column=3), proc_n.get(proc, 0),
                   bg=bg, h_align="center"); r += 1

    blank_row(ws, r, height=8); r += 1

    # ── Table 1 characteristics ────────────────────────────────────────────
    groups_in_study = meta["groups"]
    table1_data     = meta.get("table1", {})
    n_groups        = len(groups_in_study)

    # cols: B=variable label, then 2 per group (N, value), then 2 for Total
    total_n_col    = 3 + n_groups * 2
    total_stat_col = 4 + n_groups * 2
    t1_col_end     = total_stat_col

    # ensure wide-enough columns for extra group/total cols beyond F
    for col_idx in range(7, t1_col_end + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=t1_col_end)
    style_cell(ws.cell(row=r, column=2),
               "TABLE 1: PARTICIPANT CHARACTERISTICS  "
               "(N = non-missing per variable; one value per participant)",
               bg=MID_BLUE, fg=WHITE, bold=True, size=11)
    ws.row_dimensions[r].height = 20; r += 1

    # header: Variable | Group1 N | Group1 Stat | ... | Total N | Total Stat
    hdr_vals = ["Variable"]
    for g in groups_in_study:
        hdr_vals += [f"{g}\nN", f"{g}\nMean / Median / %"]
    hdr_vals += ["Total\nN", "Total\nMean / Median / %"]
    header_row(ws, r, hdr_vals, col_start=2, height=32)
    r += 1

    t1_display = [(col, lbl, stat) for col, lbl, stat in TABLE1_VARS
                  if stat != "category"]

    for i, (col, lbl, stat) in enumerate(t1_display):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20
        style_cell(ws.cell(row=r, column=2), lbl, bg=bg, fg=DARK_BLUE, bold=True)

        for j, grp in enumerate(groups_in_study):
            values = table1_data.get(grp, {}).get(col, [])
            n, val = compute_stats(values, stat)
            style_cell(ws.cell(row=r, column=3 + j * 2), n,
                       bg=bg, h_align="center")
            style_cell(ws.cell(row=r, column=4 + j * 2), val,
                       bg=bg, h_align="center")

        # Total column: combine values across all groups
        all_values = []
        for grp in groups_in_study:
            all_values.extend(table1_data.get(grp, {}).get(col, []))
        total_n, total_val = compute_stats(all_values, stat)
        style_cell(ws.cell(row=r, column=total_n_col), total_n,
                   bg=bg, h_align="center", bold=True)
        style_cell(ws.cell(row=r, column=total_stat_col), total_val,
                   bg=bg, h_align="center", bold=True)
        r += 1

    blank_row(ws, r, height=8); r += 1

    # ── Omics data availability ────────────────────────────────────────────
    section_row(ws, r, 2, 6, "OMICS DATA AVAILABILITY"); r += 1
    header_row(ws, r, ["Omics Type", "# Participants with Data"],
               col_start=2, height=20)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    r += 1

    om = meta.get("omics", {})
    omics_rows = [
        ("SomaScan (Proteomics)",      om.get("soma",    0)),
        ("Olink Plasma (Proteomics)",  om.get("olink_p", 0)),
        ("Olink Urine (Proteomics)",   om.get("olink_u", 0)),
        ("Metabolomics",               om.get("metab",   0)),
        ("Lipidomics",                 om.get("lipid",   0)),
    ]
    for i, (label, n) in enumerate(omics_rows):
        bg = LIGHT_GREEN if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20
        style_cell(ws.cell(row=r, column=2), label, bg=bg, fg=DARK_BLUE)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        style_cell(ws.cell(row=r, column=3), n, bg=bg, h_align="center")
        r += 1
    # Total (any omics) row
    ws.row_dimensions[r].height = 20
    style_cell(ws.cell(row=r, column=2), "Total (any omics)",
               bg=DARK_BLUE, fg=WHITE, bold=True, size=11)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    style_cell(ws.cell(row=r, column=3), om.get("any", 0),
               bg=DARK_BLUE, fg=WHITE, h_align="center", bold=True, size=11)
    r += 1

    blank_row(ws, r, height=8); r += 1

    # ── PB90 scRNA inventory ──────────────────────────────────────────────
    if scrna:
        scrna_groups = sorted(scrna["by_group"].keys())
        n_scrna_grps = len(scrna_groups)
        # cols: B=Cell Type, C..C+n-1=groups, C+n=Total
        scrna_col_end = 3 + n_scrna_grps   # last column index
        # Ensure column widths are set for all PB90 columns
        for col_idx in range(3, scrna_col_end + 2):
            ltr = get_column_letter(col_idx)
            if ltr not in col_widths:
                ws.column_dimensions[ltr].width = 18
        hdr_span = max(t1_col_end, scrna_col_end)
    else:
        hdr_span = t1_col_end

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=hdr_span)
    style_cell(ws.cell(row=r, column=2),
               "PB90 scRNA INVENTORY  (pb90_meta.rds — Kopah S3)",
               bg=MID_BLUE, fg=WHITE, bold=True, size=11)
    ws.row_dimensions[r].height = 20; r += 1

    if not scrna:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=hdr_span)
        style_cell(ws.cell(row=r, column=2),
                   "No PB90 scRNA data available for this study.",
                   bg=GREY, fg="888888", italic=True)
        ws.row_dimensions[r].height = 20; r += 1
    else:
        # Overview row: total participants only
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=hdr_span)
        style_cell(ws.cell(row=r, column=2),
                   f"Total participants: {scrna['n_subjects']:,}",
                   bg=LIGHT_GREEN, fg=DARK_BLUE, bold=True)
        ws.row_dimensions[r].height = 20; r += 1

        # Participants-per-group row
        ws.row_dimensions[r].height = 20
        style_cell(ws.cell(row=r, column=2), "Participants with Data",
                   bg=LIGHT_GREEN, fg=DARK_BLUE, bold=True, size=11)
        for j, grp in enumerate(scrna_groups):
            n_subj = scrna["by_group"].get(grp, {}).get("n_subjects", 0)
            style_cell(ws.cell(row=r, column=3 + j), n_subj,
                       bg=LIGHT_GREEN, h_align="center", bold=True)
        style_cell(ws.cell(row=r, column=3 + n_scrna_grps), scrna["n_subjects"],
                   bg=LIGHT_GREEN, h_align="center", bold=True)
        r += 1

        # Header: Cell Type | Group1 | Group2 | ...
        pb90_hdr = ["Cell Type"] + scrna_groups
        header_row(ws, r, pb90_hdr, col_start=2, height=20); r += 1

        # Sort cell types by overall total count, descending
        sorted_ct = sorted(scrna["cell_types"].items(),
                           key=lambda x: x[1], reverse=True)

        grp_totals = {g: 0 for g in scrna_groups}

        for i, (ct, ct_total) in enumerate(sorted_ct):
            bg = LIGHT_BLUE if i % 2 == 0 else WHITE
            ws.row_dimensions[r].height = 20
            style_cell(ws.cell(row=r, column=2), ct, bg=bg, fg=DARK_BLUE)
            for j, grp in enumerate(scrna_groups):
                n_ct_grp = (scrna["by_group"]
                            .get(grp, {})
                            .get("cell_types", {})
                            .get(ct, 0))
                grp_totals[grp] += n_ct_grp
                style_cell(ws.cell(row=r, column=3 + j), n_ct_grp,
                           bg=bg, h_align="center")
            r += 1

        # ── Total row for PB90 table ──────────────────────────────────────
        ws.row_dimensions[r].height = 20
        style_cell(ws.cell(row=r, column=2), "Total",
                   bg=DARK_BLUE, fg=WHITE, bold=True, size=11)
        for j, grp in enumerate(scrna_groups):
            style_cell(ws.cell(row=r, column=3 + j), grp_totals[grp],
                       bg=DARK_BLUE, fg=WHITE, h_align="center",
                       bold=True, size=11)
        r += 1

    # ── Data dictionary (right-side panel, separate row counter dr) ────────
    # Grouped by procedure name with an addition section for rows without procedure.
    # Each section header is the procedure name
    # (as it appears in the 'procedure' column), not the REDCap form_name.
    if data_dict and proc_groups:
        dr = 1  # independent row counter; starts at the top of the sheet

        # Title bar
        ws.merge_cells(start_row=dr, start_column=DICT_VAR_COL,
                       end_row=dr,   end_column=DICT_N_COL)
        style_cell(ws.cell(row=dr, column=DICT_VAR_COL),
                   f"DATA DICTIONARY — {study_name}",
                   bg=DARK_BLUE, fg=WHITE, bold=True, size=13)
        dr += 1
        blank_row(ws, dr, height=8); dr += 1

        var_n = meta.get("var_n", {})

        for proc_name, proc_vars in proc_groups.items():
            # Procedure sub-header (uses procedure name, not REDCap form_name)
            ws.merge_cells(start_row=dr, start_column=DICT_VAR_COL,
                           end_row=dr,   end_column=DICT_N_COL)
            n_label = len(proc_vars) if proc_vars else "no variables identified"
            style_cell(ws.cell(row=dr, column=DICT_VAR_COL),
                       f"{proc_name}  ({n_label})",
                       bg=MID_BLUE, fg=WHITE, bold=True, size=11)
            ws.row_dimensions[dr].height = 20; dr += 1

            if not proc_vars:
                # Placeholder row for procedures with no matched dict variables
                ws.merge_cells(start_row=dr, start_column=DICT_VAR_COL,
                               end_row=dr,   end_column=DICT_N_COL)
                style_cell(ws.cell(row=dr, column=DICT_VAR_COL),
                           "No clinical variables identified in data dictionary for this procedure.",
                           bg=GREY, fg="888888", italic=True, size=9)
                ws.row_dimensions[dr].height = 20; dr += 1
            else:
                # Column headers
                for col_idx, lbl in [
                    (DICT_VAR_COL,   "Variable"),
                    (DICT_LABEL_COL, "Label"),
                    (DICT_UNITS_COL, "Units"),
                    (DICT_N_COL,     "N Participants"),
                ]:
                    style_cell(ws.cell(row=dr, column=col_idx), lbl,
                               bg=LIGHT_BLUE, fg=DARK_BLUE, size=9,
                               bold=True, h_align="center")
                style_cell(ws.cell(row=dr, column=DICT_FORM_COL), "",
                           bg=LIGHT_BLUE)
                ws.row_dimensions[dr].height = 20; dr += 1

                for i, (col_name, info) in enumerate(proc_vars):
                    bg = LIGHT_GREEN if i % 2 == 0 else WHITE
                    n_participants = var_n.get((proc_name, col_name), 0)
                    style_cell(ws.cell(row=dr, column=DICT_VAR_COL),
                               col_name, bg=bg, fg=DARK_BLUE, size=9)
                    style_cell(ws.cell(row=dr, column=DICT_LABEL_COL),
                               info.get("label", ""), bg=bg, size=9)
                    style_cell(ws.cell(row=dr, column=DICT_UNITS_COL),
                               info.get("units", ""), bg=bg, size=9, h_align="center")
                    style_cell(ws.cell(row=dr, column=DICT_FORM_COL),
                               "", bg=bg)
                    style_cell(ws.cell(row=dr, column=DICT_N_COL),
                               n_participants if n_participants else "",
                               bg=bg, size=9, h_align="center")
                    dr += 1

            blank_row(ws, dr, height=8); dr += 1


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Reading dataset and computing statistics …")
    print("(This may take a minute — the dataset has ~28,000 columns.)\n")

    print("Loading data dictionary …")
    data_dict = load_data_dictionary(DATA_DICT_FILE)

    study_meta, soma_cols, olink_cols, clinical_cols, proc_group_n = load_study_metadata(DATA_FILE)

    print(f"\n  {len(study_meta)} studies found:")
    for name, m in study_meta.items():
        print(f"    {name:22s}  n_subjects={len(m['subjects']):>3d}"
              f"  procedures={len(m['procedures'])}"
              f"  study_cols={len(m['study_cols'])}")

    print("\nLoading PB90 scRNA metadata from S3 …")
    scrna_meta = load_scrna_metadata()
    if scrna_meta:
        print(f"  PB90 studies present: {', '.join(sorted(scrna_meta.keys()))}")

    print("\nBuilding workbook …")
    wb = build_workbook(study_meta, soma_cols, olink_cols, clinical_cols,
                        scrna_meta, data_dict=data_dict,
                        proc_group_n=proc_group_n)

    print(f"Saving → {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("Done.")
